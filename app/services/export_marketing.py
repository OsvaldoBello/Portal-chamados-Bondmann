"""Exportação organizada e dinâmica dos indicadores do Dashboard de Marketing.

Pedido do usuário (2026-07-31): a exportação existente (`/admin/export/csv`,
link "Exportar CSV" da barra lateral) é uma tabela ÚNICA e genérica com os
chamados brutos de qualquer setor — não reflete os indicadores AGREGADOS do
dashboard de Marketing (taxa de entrega mensal, volume produzido, origem,
ranking de solicitantes, tempo médio/atrasos, mídia regional) nem respeita o
período selecionado na tela. Esta exportação (rota própria,
`/admin/marketing/export`):

- é **organizada**: uma aba por indicador — mesma divisão das 6 abas do
  dashboard — em vez de uma tabela achatada de chamados;
- é **dinâmica**: recebe o mesmo ``periodo`` do filtro do dashboard ("all" =
  Acumulado, ou o rótulo de um mês, ex. "JUL/26" — mesmo valor do atributo
  `data-filter` dos "pills") e filtra os dados ANTES de montar a planilha; o
  botão "Exportar" do dashboard manda automaticamente o filtro que está ativo
  na tela no momento do clique (`app/static/js/admin_marketing.js`).

Consome o MESMO dict que `AdminRepo.mkt_dashboard_data` monta para a tela —
os números da planilha nunca divergem dos números mostrados no navegador (sem
segunda fonte de verdade / segunda query)."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_HEADER_FILL = PatternFill(start_color="FF1A3A6B", end_color="FF1A3A6B", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFFFF", bold=True)
_TITLE_FONT = Font(bold=True, size=13, color="FF1A3A6B")


def _filtrar_por_periodo(mkt_data: dict[str, Any], periodo: str) -> dict[str, Any]:
    """``periodo`` = "all" (acumulado) ou o rótulo de um mês (ex. "JUL/26").

    Filtra `monthly`/`atrasosData`/`deptByMonth` pelo mesmo rótulo usado nos
    "pills" do dashboard. `midia` é indexado só por nome de mês, sem ano
    (mesma limitação de `admin_marketing.js::updateMidiaInsights` — casamento
    pelo prefixo do rótulo, ex. "JUL/26" → "Jul")."""
    monthly = mkt_data.get("monthly", [])
    if periodo and periodo != "all":
        monthly = [m for m in monthly if m["label"] == periodo]
    labels = {m["label"] for m in monthly}

    atrasos = [a for a in mkt_data.get("atrasosData", []) if a["mes"] in labels]
    dept_by_month = {k: v for k, v in mkt_data.get("deptByMonth", {}).items() if k in labels}

    midia = mkt_data.get("midia") or {}
    midia_meses = midia.get("meses", [])
    if periodo and periodo != "all":
        prefixo = periodo.split("/", 1)[0].capitalize()
        idxs = [i for i, m in enumerate(midia_meses) if m == prefixo]
    else:
        idxs = list(range(len(midia_meses)))
    midia_filtrada = {
        chave: [midia.get(chave, [])[i] for i in idxs]
        for chave in ("investimento", "regioes", "descontinuidades", "aderencias")
    }
    midia_filtrada["meses"] = [midia_meses[i] for i in idxs]

    return {
        "monthly": monthly, "atrasosData": atrasos,
        "deptByMonth": dept_by_month, "midia": midia_filtrada,
    }


def _titulo(ws: Worksheet, texto: str, n_cols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=texto)
    c.font = _TITLE_FONT
    ws.row_dimensions[1].height = 22


def _cabecalho(ws: Worksheet, row: int, colunas: list[str]) -> None:
    for i, texto in enumerate(colunas, start=1):
        c = ws.cell(row=row, column=i, value=texto)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT


def _autofit(ws: Worksheet, larguras: list[int]) -> None:
    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura


def _aba_resumo(wb: Workbook, dados: dict[str, Any], periodo: str) -> None:
    monthly = dados["monthly"]
    ws = wb.create_sheet("Resumo")
    _titulo(ws, f"Resumo — {'Acumulado' if periodo == 'all' else periodo}", 2)

    total = sum(m["total"] for m in monthly)
    conc = sum(m["concluidas"] for m in monthly)
    em_andamento = sum(m["em_andamento"] for m in monthly)
    abertas = sum(m["abertas"] for m in monthly)
    volume = sum(m["volume"] for m in monthly)
    mkt_orig = sum(m["mkt_orig"] for m in monthly)
    atrasos = len(dados["atrasosData"])
    tempo_medio = (
        sum(m["tempo_medio"] * m["concluidas"] for m in monthly) / conc if conc else 0.0
    )
    agg: dict[str, int] = {}
    for por_setor in dados["deptByMonth"].values():
        for setor, qtd in por_setor.items():
            agg[setor] = agg.get(setor, 0) + qtd
    top_setor = max(agg.items(), key=lambda kv: kv[1]) if agg else ("—", 0)

    linhas = [
        ("Total de demandas", total),
        ("Concluídas", conc),
        ("% concluídas", round(100 * conc / total, 1) if total else 0.0),
        ("Em andamento", em_andamento),
        ("Abertas", abertas),
        ("Volume produzido (peças/cards)", volume),
        ("Tempo médio de entrega (dias, só concluídas)", round(tempo_medio, 1)),
        ("Atrasos > 5 dias", atrasos),
        ("% atrasos sobre o total", round(100 * atrasos / total, 1) if total else 0.0),
        ("Demandas com origem Marketing", mkt_orig),
        ("% origem Marketing", round(100 * mkt_orig / total, 1) if total else 0.0),
        ("Maior setor solicitante", top_setor[0]),
        ("Demandas do maior solicitante", top_setor[1]),
    ]
    _cabecalho(ws, 3, ["Indicador", "Valor"])
    for i, (label, valor) in enumerate(linhas, start=4):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=valor)
    _autofit(ws, [42, 16])


def _aba_volume_mensal(wb: Workbook, dados: dict[str, Any]) -> None:
    ws = wb.create_sheet("Volume Mensal")
    colunas = [
        "Mês", "Total", "Concluídas", "Em andamento", "Abertas", "Volume produzido",
        "Origem Marketing", "Origem Solicitação", "Tempo médio (d)", "Atrasos > 5d",
    ]
    _titulo(ws, "Taxa de Entrega e Volume Produzido — Histórico Mensal", len(colunas))
    _cabecalho(ws, 3, colunas)
    for i, m in enumerate(dados["monthly"], start=4):
        valores = [
            m["label"], m["total"], m["concluidas"], m["em_andamento"], m["abertas"],
            m["volume"], m["mkt_orig"], m["sol_orig"], m["tempo_medio"], m["atrasos"],
        ]
        for c, v in enumerate(valores, start=1):
            ws.cell(row=i, column=c, value=v)
    _autofit(ws, [10, 8, 11, 13, 9, 16, 16, 17, 15, 13])


def _aba_solicitantes(wb: Workbook, dados: dict[str, Any]) -> None:
    ws = wb.create_sheet("Solicitantes")
    _titulo(ws, "Demandas por Setor Solicitante — Ranking", 2)
    agg: dict[str, int] = {}
    for por_setor in dados["deptByMonth"].values():
        for setor, qtd in por_setor.items():
            agg[setor] = agg.get(setor, 0) + qtd
    _cabecalho(ws, 3, ["Setor solicitante", "Demandas"])
    for i, (setor, qtd) in enumerate(sorted(agg.items(), key=lambda kv: kv[1], reverse=True), start=4):
        ws.cell(row=i, column=1, value=setor)
        ws.cell(row=i, column=2, value=qtd)
    _autofit(ws, [32, 12])


def _aba_atrasos(wb: Workbook, dados: dict[str, Any]) -> None:
    ws = wb.create_sheet("Tempo e Atrasos")
    colunas = ["Demanda", "Mês", "Dias em aberto", "Causa do atraso"]
    _titulo(ws, "Demandas com Atraso > 5 dias", len(colunas))
    _cabecalho(ws, 3, colunas)
    for i, a in enumerate(dados["atrasosData"], start=4):
        ws.cell(row=i, column=1, value=a["nome"])
        ws.cell(row=i, column=2, value=a["mes"])
        ws.cell(row=i, column=3, value=a["dias"])
        ws.cell(row=i, column=4, value=a["causa"])
    _autofit(ws, [40, 10, 15, 32])


def _aba_midia_regional(wb: Workbook, dados: dict[str, Any]) -> None:
    ws = wb.create_sheet("Mídia Regional")
    colunas = ["Mês", "Investimento (R$)", "Regiões ativas", "Descontinuidades", "Aderências"]
    _titulo(ws, "Mídia Regional — Investimento por Região", len(colunas))
    _cabecalho(ws, 3, colunas)
    midia = dados["midia"]
    for i, mes in enumerate(midia["meses"], start=4):
        idx = i - 4
        ws.cell(row=i, column=1, value=mes)
        ws.cell(row=i, column=2, value=midia["investimento"][idx])
        ws.cell(row=i, column=3, value=midia["regioes"][idx])
        ws.cell(row=i, column=4, value=midia["descontinuidades"][idx])
        ws.cell(row=i, column=5, value=midia["aderencias"][idx])
    _autofit(ws, [10, 18, 15, 17, 13])


def gerar_workbook(mkt_data: dict[str, Any], periodo: str = "all") -> BytesIO:
    """``mkt_data`` = retorno de `AdminRepo.mkt_dashboard_data`; ``periodo`` =
    "all" ou o rótulo de um mês (mesmo valor do filtro da tela). Devolve os
    bytes do `.xlsx` prontos para a resposta HTTP."""
    dados = _filtrar_por_periodo(mkt_data, periodo)
    wb = Workbook()
    wb.remove(wb.active)  # aba padrão vazia — cada indicador ganha a sua
    _aba_resumo(wb, dados, periodo)
    _aba_volume_mensal(wb, dados)
    _aba_solicitantes(wb, dados)
    _aba_atrasos(wb, dados)
    _aba_midia_regional(wb, dados)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
