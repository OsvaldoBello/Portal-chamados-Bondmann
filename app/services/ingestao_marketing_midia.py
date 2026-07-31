"""Ingestão da planilha "Investimento por Região" → `marketing_midia_regional`.

O time de Marketing recebe mensalmente, de uma agência externa, uma planilha
com **uma aba por mês** ("Março 2026", "Abril 2026", ...) listando os
representantes/regiões de veiculação. Este módulo interpreta essa planilha de
forma DETERMINÍSTICA (openpyxl, sem LLM) e devolve os 4 indicadores da aba
"Mídia Regional" do dashboard (mesmos campos de `AdminRepo.upsert_marketing_midia_regional`):

- **Investimento**: soma da coluna cujo cabeçalho começa com ``"5."`` (ex.:
  "5.Valor mensal do investimento (vendedor + Bondmann):") — o valor REAL
  reportado por representante, não o "disponível" — dividida por 2 (regra
  passada pelo usuário 2026-07-30: a planilha soma vendedor + Bondmann, o
  indicador do dashboard mostra só a metade Bondmann).
- **Regiões ativas**: maior número da coluna A (a sequência 1..N dos
  representantes) que estiver com preenchimento **verde sólido**
  (``FF00FF00``) — reps sem preenchimento verde não contam.
- **Descontinuidades**/**Aderências**: contagem de células da coluna cujo
  cabeçalho começa com ``"6."`` e contém "obs" (ex.: "6.Observações") cujo
  texto começa com "Saiu"/"Entrou" — o template combinado com o usuário
  2026-07-30 ("Saiu em 'dia' de 'mês'.", "Entrou em 'dia' de 'mês'.") e usado
  de fato nos meses mais recentes da planilha real (conferido linha a linha:
  bate 100% com jun/2026). Meses mais antigos têm observações em texto livre
  ("Início em...", "Retornou em...", "Promovido a supervisor...") que NÃO
  seguem esse template — ficam de fora da contagem de propósito: o parser
  nunca adivinha sinônimo, só o texto combinado conta.

Layout validado linha a linha contra `Indicadores (1).xlsx` (aba
`MidiaCompartilhada`, valores de referência já calculados à mão pelo
Marketing p/ jan-jun/2026): Investimento e Regiões Ativas batem exatamente
nos 6 meses; Descontinuidades/Aderências batem nos meses em que a coluna de
observações já seguia o template combinado (ex. jun/2026 bate 100% — 1
descontinuidade, 3 aderências). Meses antigos (nov/2024 a jul/2025) não têm
coluna de observações nenhuma (o template só passou a existir depois) —
nesses meses `descontinuidades`/`aderencias` saem como 0 (nunca inventa dado
que não está na planilha).

Abas auxiliares ("Copy of ...", resumos, gráficos) e qualquer aba cujo nome
não bata com o padrão "‹Mês em português› ‹Ano de 4 dígitos›" são ignoradas.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from datetime import date
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO
from typing import Any

_MESES_PT: dict[str, int] = {
    "janeiro": 1, "fevereiro": 2, "marco": 3, "abril": 4, "maio": 5, "junho": 6,
    "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
}

# Verde sólido usado pela planilha para marcar representante/região ativo na
# coluna A. `openpyxl` devolve o RGB com prefixo de alpha (ARGB).
_VERDE_ATIVO = "FF00FF00"

_RE_ABA_MES = re.compile(r"^\s*([A-Za-zÀ-ÿ]+)\s+(\d{4})\s*$")
_RE_SAIU = re.compile(r"^\s*saiu\b", re.IGNORECASE)
_RE_ENTROU = re.compile(r"^\s*entrou\b", re.IGNORECASE)


def _sem_acento(texto: str) -> str:
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _texto(valor: Any) -> str:
    return str(valor).strip() if valor is not None else ""


def mes_da_aba(nome_aba: str) -> date | None:
    """"Março 2026" → ``date(2026, 3, 1)``; ``None`` se a aba não for um mês
    (ex.: "Copy of Fevereiro 2025", "Indicadores", "Gráficos")."""
    m = _RE_ABA_MES.match(nome_aba)
    if not m:
        return None
    mes_num = _MESES_PT.get(_sem_acento(m.group(1)).lower())
    if not mes_num:
        return None
    return date(int(m.group(2)), mes_num, 1)


def _achar_cabecalho(ws) -> tuple[int, int | None, int | None] | None:
    """Acha a linha de cabeçalho (procura "Representante" nas 3 primeiras
    linhas — algumas abas têm uma linha de título acima do cabeçalho) e
    devolve ``(linha, coluna_investimento, coluna_observacoes)``.

    As colunas são achadas pelo TEXTO do cabeçalho (``"5."``/``"6."`` +
    "obs"), nunca por posição fixa — o layout mudou de mês para mês (colunas
    de leads/Neat foram adicionadas depois, 2025)."""
    for r in range(1, 4):
        max_col = ws.max_column or 1
        cols = {c: _texto(ws.cell(row=r, column=c).value) for c in range(1, max_col + 1)}
        if not any(v.lower().startswith("representante") for v in cols.values()):
            continue
        col_investimento = next(
            (c for c, v in cols.items() if v.lower().startswith("5.")), None
        )
        col_obs = next(
            (c for c, v in cols.items() if v.lower().startswith("6.") and "obs" in v.lower()),
            None,
        )
        return r, col_investimento, col_obs
    return None


def _eh_verde(cell) -> bool:
    fill = cell.fill
    fg = fill.fgColor if fill else None
    return bool(fill and fill.patternType == "solid" and fg and fg.rgb == _VERDE_ATIVO)


def _para_decimal(valor: Any) -> Decimal | None:
    """Célula → ``Decimal`` (nunca ``float`` direto: soma de dezenas de
    centavos em ponto flutuante binário produz ruído tipo 4057.0149999999996
    que arredonda para baixo do esperado — ``Decimal`` a partir do texto/valor
    original evita isso)."""
    if isinstance(valor, (int, float)):
        return Decimal(str(valor))
    if isinstance(valor, str) and valor.strip():
        try:
            return Decimal(valor.strip().replace(",", "."))
        except Exception:  # noqa: BLE001 — texto não numérico na célula
            return None
    return None


@dataclass
class RegistroMidia:
    mes: date
    investimento: float
    regioes: int
    descontinuidades: int
    aderencias: int


def ler_aba(ws) -> tuple[float, int, int, int] | None:
    """Uma aba de mês → ``(investimento, regioes, descontinuidades, aderencias)``.

    ``None`` se a aba não tiver o cabeçalho esperado (não é uma aba de dados
    de representantes — ex.: aba de resumo)."""
    cabecalho = _achar_cabecalho(ws)
    if cabecalho is None:
        return None
    linha_cabecalho, col_investimento, col_obs = cabecalho

    regioes_ativas = 0
    investimento_total = Decimal("0")
    descontinuidades = 0
    aderencias = 0

    for r in range(linha_cabecalho + 1, ws.max_row + 1):
        id_cell = ws.cell(row=r, column=1)
        if not isinstance(id_cell.value, (int, float)):
            continue  # pula linha "TOTAL"/em branco/fora do bloco de dados
        if _eh_verde(id_cell):
            regioes_ativas = max(regioes_ativas, int(id_cell.value))
        if col_investimento:
            v = _para_decimal(ws.cell(row=r, column=col_investimento).value)
            if v is not None:
                investimento_total += v
        if col_obs:
            obs = _texto(ws.cell(row=r, column=col_obs).value)
            if _RE_SAIU.match(obs):
                descontinuidades += 1
            elif _RE_ENTROU.match(obs):
                aderencias += 1

    if col_investimento:
        investimento = float(
            (investimento_total / 2).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        )
    else:
        investimento = 0.0
    return investimento, regioes_ativas, descontinuidades, aderencias


def parse_planilha(wb) -> list[RegistroMidia]:
    """Workbook aberto → lista de registros mensais (1 por aba de mês
    reconhecida), ordenada por mês. Abas que não sejam "‹Mês› ‹Ano›" (abas
    "Copy of ...", resumos etc.) ou sem o cabeçalho esperado são ignoradas
    silenciosamente — planilha real tem abas assim."""
    registros: list[RegistroMidia] = []
    for nome_aba in wb.sheetnames:
        if nome_aba.strip().lower().startswith("copy of"):
            continue
        mes = mes_da_aba(nome_aba)
        if mes is None:
            continue
        dados = ler_aba(wb[nome_aba])
        if dados is None:
            continue
        investimento, regioes, descontinuidades, aderencias = dados
        registros.append(
            RegistroMidia(
                mes=mes, investimento=investimento, regioes=regioes,
                descontinuidades=descontinuidades, aderencias=aderencias,
            )
        )
    registros.sort(key=lambda r: r.mes)
    return registros


def parse_bytes(conteudo: bytes) -> list[RegistroMidia]:
    """Interpreta a planilha EM MEMÓRIA (determinístico, sem LLM) — ponto de
    entrada usado pela rota web.

    **Sem** ``read_only=True``: o modo somente-leitura do openpyxl não
    resolve `max_row`/`max_column` sem uma varredura prévia (voltam `None`) e
    `_achar_cabecalho`/`ler_aba` dependem de acesso aleatório por
    linha/coluna — o arquivo é pequeno o bastante (~1MB, dezenas de abas)
    para o modo normal ser rápido o suficiente."""
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(conteudo), data_only=True)
    return parse_planilha(wb)
