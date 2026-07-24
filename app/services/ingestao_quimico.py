"""Ingestão da base de conhecimento do Dpto Químico → tabelas ``base_quimico_*`` (F4).

plano_md_mestre_IA.md, Seção 3.3: a base (planilha de 16 abas + PDF de fichas
técnicas) é SIGILOSA e **nunca entra no repositório**. Este módulo **interpreta
os arquivos de forma DETERMINÍSTICA (openpyxl/pypdf, sem LLM)** e faz *upsert*
no banco — as quantidades das formulações jamais passam por um modelo (Regra de
Ouro #4). É consumido por dois pontos de entrada:

- **CLI** (`scripts/ingestao_base_quimico.py`): o dev/gestor roda localmente
  apontando os caminhos dos arquivos.
- **Painel web** (`/admin/base-quimico`, F4): o químico sobe a planilha/PDF
  pela tela; a rota chama :func:`ingerir_conn` na conexão administrativa.

O que vai para onde (Seção 3.3 do plano IA):
- ``Base_IA_Produtos``       → ``base_quimico_produtos`` (componentes SEM proporção)
- ``Base_IA_Materias_Primas``→ ``base_quimico_materias_primas``
- ``Diagnostico_Ocorrencias``→ ``base_quimico_playbooks`` (tipo DIAGNOSTICO)
- ``Perguntas_Investigacao`` → ``base_quimico_playbooks`` (tipo PERGUNTA_INVESTIGACAO)
- ``Regras_Sigilo_Resposta`` → ``base_quimico_playbooks`` (tipo REGRA_SIGILO)
- PDF de fichas (por produto)→ ``base_quimico_fichas``
- ``Base_IA_Componentes``    → ``base_quimico_formulacoes`` (quantidades — só no
  banco; sem GRANT ao role ``ia_worker``; nenhum passe de modelo lê — C7)

As abas-modelo em preenchimento (Compatibilidade_Materiais, Parametros_Controle,
RCA_6M etc.) não são ingeridas na v1 (Seção 3.3 — "não bloqueiam a v1").
"""

from __future__ import annotations

import asyncio
import json
import re
import unicodedata
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

# ---------------------------------------------------------------------------
# Parsing da planilha (funções puras — testáveis sem arquivo real)
# ---------------------------------------------------------------------------


def _texto(valor: Any) -> str | None:
    """Célula → str limpa (None para vazio)."""
    if valor is None:
        return None
    s = str(valor).strip()
    return s or None


def _linhas(ws) -> list[dict[str, Any]]:
    """Aba → lista de dicts {cabeçalho: valor}, ignorando linhas vazias."""
    it = ws.iter_rows(values_only=True)
    try:
        cabecalhos = [(_texto(c) or f"col{i}") for i, c in enumerate(next(it))]
    except StopIteration:
        return []
    out = []
    for row in it:
        if not any(v is not None and str(v).strip() for v in row):
            continue
        out.append({cabecalhos[i]: row[i] if i < len(row) else None for i in range(len(cabecalhos))})
    return out


def ler_produtos(ws) -> list[dict[str, Any]]:
    """Aba Base_IA_Produtos → linhas de ``base_quimico_produtos``.

    ``componentes`` vem da coluna "Componentes principais (sem proporção)" —
    NOMES separados por ';', nunca quantidades."""
    produtos = []
    for r in _linhas(ws):
        chave = _texto(r.get("Chave Produto"))
        nome = _texto(r.get("Produto"))
        if not chave or not nome:
            continue
        componentes = [
            c.strip() for c in (_texto(r.get("Componentes principais (sem proporção)")) or "").split(";") if c.strip()
        ]
        palavras = [
            p.strip() for p in (_texto(r.get("Palavras-chave para busca da IA")) or "").split() if p.strip()
        ]
        produtos.append(
            {
                "chave_produto": chave,
                "segmento": _texto(r.get("Segmento")),
                "codigo_produto": _texto(r.get("Código Produto")),
                "nome": nome,
                "nome_normalizado": _texto(r.get("Nome normalizado")),
                "aplicacao": _texto(r.get("Descrição / aplicação")),
                "familia_tecnica": _texto(r.get("Família técnica sugerida")),
                "tipo_uso": _texto(r.get("Tipo de uso sugerido")),
                "componentes": componentes,
                "palavras_chave": palavras,
                "orientacao": _texto(r.get("Orientação para resposta da IA")),
            }
        )
    return produtos


def ler_materias_primas(ws) -> list[dict[str, Any]]:
    """Aba Base_IA_Materias_Primas → linhas de ``base_quimico_materias_primas``."""
    mps = []
    for r in _linhas(ws):
        codigo = _texto(r.get("Código MP"))
        nome = _texto(r.get("Nome classificado"))
        if not codigo or not nome:
            continue
        mps.append(
            {
                "codigo_mp": codigo,
                "nome": nome,
                "formula_quimica": _texto(r.get("Fórmula química")),
                "utilizacoes": _texto(r.get("Principais utilizações")),
            }
        )
    return mps


def ler_formulacoes(ws) -> list[dict[str, Any]]:
    """Aba Base_IA_Componentes → ``base_quimico_formulacoes`` (quantidades).

    Estas linhas são "Confidencial / formulação": ficam SÓ no banco, fora do
    alcance do role ``ia_worker`` (C7) — nunca em contexto de modelo."""
    linhas = []
    for r in _linhas(ws):
        chave = _texto(r.get("Chave Produto"))
        componente = _texto(r.get("Matéria-prima / componente"))
        ordem = r.get("Ordem")
        if not chave or not componente or ordem is None:
            continue
        try:
            quantidade = float(str(r.get("Quantidade informada")).replace(",", "."))
        except (TypeError, ValueError):
            quantidade = None
        linhas.append(
            {
                "chave_produto": chave,
                "ordem": int(float(ordem)),
                "codigo_mp": _texto(r.get("Código MP/Componente")),
                "componente": componente,
                "quantidade": quantidade,
                "funcao": _texto(r.get("Função/observação extraída")),
            }
        )
    return linhas


# (tipo, coluna-chave) de cada aba de playbook — as demais colunas viram `dados`.
_PLAYBOOKS = (
    ("DIAGNOSTICO", "Diagnostico_Ocorrencias", "Sintoma / problema relatado"),
    ("PERGUNTA_INVESTIGACAO", "Perguntas_Investigacao", "Cenário / sintoma"),
    ("REGRA_SIGILO", "Regras_Sigilo_Resposta", "Tipo de informação solicitada"),
)


def ler_playbooks(wb) -> list[dict[str, Any]]:
    """Abas de diagnóstico/perguntas/regras → ``base_quimico_playbooks``."""
    out = []
    for tipo, aba, coluna_chave in _PLAYBOOKS:
        if aba not in wb.sheetnames:
            print(f"  AVISO: aba '{aba}' ausente na planilha — pulada.")
            continue
        for r in _linhas(wb[aba]):
            sintoma = _texto(r.get(coluna_chave))
            if not sintoma:
                continue
            dados = {
                k: _texto(v) for k, v in r.items() if k != coluna_chave and _texto(v) is not None
            }
            out.append({"tipo": tipo, "sintoma": sintoma, "dados": dados})
    return out


# ---------------------------------------------------------------------------
# Fatiamento do PDF de fichas por produto (função pura sobre lista de páginas)
# ---------------------------------------------------------------------------


def _normalizar(texto: str) -> str:
    """Caixa alta sem acento — comparação robusta de nomes de produto."""
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in nfkd if not unicodedata.combining(c)).upper()


def _alnum(texto: str) -> str:
    """Só letras/dígitos, caixa alta, sem acento — "WAY 45 - B" == "WAY45 - B",
    "D.F.D." == "DFD", ">> LB10 H" == "LB 10 H" (grafias do PDF × planilha)."""
    return re.sub(r"[^A-Z0-9]", "", _normalizar(texto))


def _padrao_flexivel(nome: str) -> re.Pattern[str]:
    """Regex do nome tolerante a espaços/pontos/hífens internos, com limite de
    palavra nas pontas ("26" não casa com "26000"; "OX" não casa com "TOXICO")."""
    partes = [re.escape(c) for c in _alnum(nome)]
    return re.compile(r"(?<![A-Z0-9])" + r"[\s.\-–—]*".join(partes) + r"(?![A-Z0-9])")


# Nome na planilha → nome(s) da(s) ficha(s) correspondente(s) no PDF.
# Conferido com o gestor/químico em 2026-07-24 (páginas citadas por eles): as
# "BASE *" são vendidas com outro(s) nome(s) comercial(is); BASE BRIL e BASE
# CONCENTRADO têm DUAS fichas cada (mesmas características, nomes/descrições
# distintos). LIMPTEC/SABOLIQ: o PDF traz a variante "INCOLOR/SEM AROMA" como
# ficha própria — anexada ao produto principal. PROTEC (não há "BASE PROTEC" na
# planilha) é o mesmo caso: o produto PROTEC concentra as fichas de PROTEC e
# PROTETIVO (mesma base). DESINFETANTE LAVANDA: ver CORRECOES_NOME_PRODUTO.
ALIASES_FICHAS: dict[str, tuple[str, ...]] = {
    "DESINFETANTE FEMME": ("FEMME",),               # pág 30
    "DESINFETANTE FLORAL": ("FLORAL",),             # pág 31
    "DESINFETANTE LAVANDA": ("LAVANDA",),           # pág 36 (S101 — ver correção)
    "BASE OX": ("OX",),                             # pág 50 — mesma ficha do OX
    "BASE SNAP": ("SNAP",),                         # pág 61 — mesma ficha do SNAP
    "BASE BRIL": ("BRIL", "GRAXCAR II"),            # págs 14 + 35
    "BASE CONCENTRADO": ("CONCENTRADO", "SHAMP"),   # págs 20 + 59
    "PROTEC": ("PROTETIVO",),                       # pág 53 (âncora) + 54 (PROTETIVO)
    "LIMPTEC 100": ("LIMPTEC 100 INCOLOR",),        # pág 41 (variante) + 42
    "SABOLIQ": ("SABOLIQ SEM AROMA INCOLOR",),      # pág 56 (variante) + 57
}

# Correção de nome preenchido errado na planilha (chave_produto → nome correto).
# 2026-07-24: o químico rotulou o S101 como "DESINFETANTE FLORAL", mas é o
# LAVANDA (S025 é o FLORAL de fato). Corrigido aqui até o gestor arrumar o
# arquivo-mestre; sem isto haveria dois "DESINFETANTE FLORAL" e o LAVANDA
# ficaria sem ficha. A chave_produto (PK) é preservada — só o nome muda.
CORRECOES_NOME_PRODUTO: dict[str, str] = {
    "BD CLEAN|S101|DESINFETANTE FLORAL|L69|013": "DESINFETANTE LAVANDA",
}

# Produtos comprados PRONTOS de terceiros: sem formulação própria (por isso
# fora da planilha de formulações), mas têm ficha técnica no PDF e aparecem no
# dropdown "Produto" do formulário de abertura. Ingeridos como produto + ficha,
# SEM componentes/formulação — não há segredo a proteger (C7 não se aplica).
# Confirmado com o químico 2026-07-24. O nome casa 1:1 com o dropdown.
PRODUTOS_SO_FICHA: tuple[str, ...] = (
    "AW-B 32", "AW-B 46", "AW-B 68",
    "LW-B 32", "LW-B 46", "LW-B 68",
)

# Fichas do PDF cujo produto NÃO está na planilha e isso é ESPERADO (forma
# alfanumérica da âncora → motivo). Mantém o aviso limpo: só reporta como
# "inesperado" o que ninguém previu. 2026-07-24 (químico).
FICHAS_ESPERADAS_SEM_PRODUTO: dict[str, str] = {
    "ADITIVO1090": "aditivo — gestor vai adicionar à planilha depois",
    "ADITIVOANTIINCRUSTANTE": "aditivo — gestor vai adicionar à planilha depois",
    "LB20H": "produto não liberado para venda — ignorar",
}


def aplicar_correcoes_nome(
    produtos: list[dict[str, Any]], correcoes: dict[str, str] = CORRECOES_NOME_PRODUTO
) -> list[dict[str, Any]]:
    """Corrige nomes preenchidos errado na planilha (por chave_produto), in-place."""
    for p in produtos:
        novo = correcoes.get(p["chave_produto"])
        if novo:
            p["nome"] = novo
    return produtos


def produtos_so_ficha(nomes: tuple[str, ...] = PRODUTOS_SO_FICHA) -> list[dict[str, Any]]:
    """Linhas mínimas de ``base_quimico_produtos`` para produtos comprados
    prontos (só ficha, sem formulação — ver ``PRODUTOS_SO_FICHA``). Sem
    componentes/quantidades; a ficha técnica carrega toda a informação."""
    return [
        {
            "chave_produto": f"FICHA|{nome}",  # chave sintética (não colide com a planilha)
            "segmento": None,
            "codigo_produto": None,
            "nome": nome,
            "nome_normalizado": None,
            "aplicacao": None,
            "familia_tecnica": None,
            "tipo_uso": None,
            "componentes": [],
            "palavras_chave": [],
            "orientacao": None,
        }
        for nome in nomes
    ]

# Âncoras de SEÇÃO das fichas (não são nome de produto) — forma alfanumérica.
_SECOES_FICHA = frozenset(
    {
        "APLICACAO",
        "INSTRUCOESDEUSO",
        "DADOSTECNICOS",
        "CARACTERISTICAS",
        "VANTAGENSEBENEFICIOS",
        "INFORMACOESDESEGURANCA",
    }
)


def fatiar_fichas(
    paginas: list[str],
    nomes_produtos: list[str],
    aliases: dict[str, tuple[str, ...]] = ALIASES_FICHAS,
) -> tuple[dict[str, str], list[tuple[int, str]]]:
    """Agrupa o texto das páginas do PDF por produto — **1 página = 1 ficha**.

    Revisão 2026-07-24 (validada página a página com o gestor): toda ficha traz
    o nome do produto numa linha-âncora ``>> NOME`` — não existem fichas de
    múltiplas páginas neste PDF. Critério:

    1. **Âncora ``>> NOME``** (comparação alfanumérica — tolera "LB10 H" ×
       "LB 10 H", "WAY 45 - B" × "WAY45 - B") decide SOZINHA. Citações a outros
       produtos no corpo ("...utilizando DEGRAX 25 e ADITIVO 967") ou palavras
       comuns ("produto concentrado") NÃO roubam a página.
    2. Página com âncora de produto DESCONHECIDO da planilha (ex.: ADITIVO
       1090, AW-B 32) é reportada — nunca anexada a outro produto.
    3. Página sem âncora nenhuma: fallback pelo corpo (regex flexível, nome
       mais longo vence, limite de palavra); sem match, é reportada.

    ``aliases`` mapeia o nome da planilha para o(s) nome(s) da ficha no PDF —
    um produto pode receber mais de uma ficha (concatenadas) e uma ficha pode
    pertencer a mais de um produto (ex.: OX e BASE OX).

    Retorna ``({nome_planilha: texto}, [(página, âncora_desconhecida)])``.
    """
    # nome de busca (como aparece no PDF) → donos (nomes da planilha)
    donos_por_busca: dict[str, list[str]] = {}
    for nome in nomes_produtos:
        donos_por_busca.setdefault(nome, []).append(nome)
        for apelido in aliases.get(nome, ()):
            donos = donos_por_busca.setdefault(apelido, [])
            if nome not in donos:
                donos.append(nome)
    por_alnum: dict[str, str] = {}
    for busca in donos_por_busca:
        por_alnum.setdefault(_alnum(busca), busca)
    padroes = [
        (busca, _padrao_flexivel(busca))
        for busca in sorted(donos_por_busca, key=len, reverse=True)
    ]

    fichas_busca: dict[str, list[str]] = {}
    desconhecidas: list[tuple[int, str]] = []
    for i, texto in enumerate(paginas):
        norm = _normalizar(texto or "")
        ancoras = [
            a
            for a in (_alnum(m.group(1)) for m in re.finditer(r">>\s*([^\n]+)", norm))
            if a and a not in _SECOES_FICHA
        ]
        dono = next((por_alnum[a] for a in ancoras if a in por_alnum), None)
        if dono is None and not ancoras:
            # Página sem âncora de produto: fallback pelo corpo do texto.
            dono = next((busca for busca, padrao in padroes if padrao.search(norm)), None)
        if dono is None:
            # Ficha de produto fora da planilha (ou página não identificada):
            # reportar para o gestor — jamais anexar a outro produto.
            desconhecidas.append((i + 1, ancoras[0] if ancoras else ""))
            continue
        fichas_busca.setdefault(dono, []).append((texto or "").strip())

    fichas: dict[str, list[str]] = {}
    for busca, partes in fichas_busca.items():
        for dono_planilha in donos_por_busca[busca]:
            fichas.setdefault(dono_planilha, []).extend(partes)
    return {nome: "\n\n".join(partes) for nome, partes in fichas.items()}, desconhecidas


# ---------------------------------------------------------------------------
# Parse dos arquivos em memória (reutilizável pela CLI e pelo painel web)
# ---------------------------------------------------------------------------


@dataclass
class ParsedBase:
    """Resultado do parse determinístico dos arquivos (sem banco, sem LLM)."""

    produtos: list[dict[str, Any]]
    mps: list[dict[str, Any]]
    playbooks: list[dict[str, Any]]
    formulacoes: list[dict[str, Any]]
    fichas_por_nome: dict[str, str]
    fichas_sem_produto: list[tuple[int, str]] = field(default_factory=list)


@dataclass
class RelatorioIngestao:
    """Resumo de uma ingestão — insumo do relatório da CLI e do painel web."""

    contagens: dict[str, int]
    fichas_esperadas: list[tuple[int, str, str]]  # (página, âncora, motivo)
    fichas_inesperadas: list[tuple[int, str]]  # (página, âncora)
    produtos_ausentes: list[str]  # nomes no banco que não vieram na planilha
    produtos_removidos: int  # quantos foram removidos (só se remover_ausentes)


def parse_planilha(wb) -> tuple[list, list, list, list]:
    """Workbook aberto → (produtos, mps, formulacoes, playbooks). As correções
    do químico e os produtos-só-ficha já entram aqui (fonte única)."""
    produtos = aplicar_correcoes_nome(ler_produtos(wb["Base_IA_Produtos"])) + produtos_so_ficha()
    mps = ler_materias_primas(wb["Base_IA_Materias_Primas"])
    formulacoes = ler_formulacoes(wb["Base_IA_Componentes"])
    playbooks = ler_playbooks(wb)
    return produtos, mps, formulacoes, playbooks


def parse_arquivos(planilha_bytes: bytes, pdf_bytes: bytes | None = None) -> ParsedBase:
    """Interpreta a planilha (+PDF) EM MEMÓRIA — determinístico, sem LLM.

    É a "interpretação do arquivo": nada de modelo de IA (as quantidades das
    formulações nunca passam por um LLM — Regra de Ouro #4). Levanta em arquivo
    inválido/aba faltando (o chamador trata)."""
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(planilha_bytes), read_only=True, data_only=True)
    produtos, mps, formulacoes, playbooks = parse_planilha(wb)
    fichas: dict[str, str] = {}
    sem_produto: list[tuple[int, str]] = []
    if pdf_bytes:
        from pypdf import PdfReader

        reader = PdfReader(BytesIO(pdf_bytes))
        paginas = [(page.extract_text() or "") for page in reader.pages]
        fichas, sem_produto = fatiar_fichas(paginas, [p["nome"] for p in produtos])
    return ParsedBase(produtos, mps, playbooks, formulacoes, fichas, sem_produto)


def _classificar_sem_produto(
    sem_produto: list[tuple[int, str]],
) -> tuple[list[tuple[int, str, str]], list[tuple[int, str]]]:
    """Fichas sem produto → (esperadas com motivo, inesperadas)."""
    esperadas = [
        (p, a, FICHAS_ESPERADAS_SEM_PRODUTO[a])
        for p, a in sem_produto
        if a in FICHAS_ESPERADAS_SEM_PRODUTO
    ]
    inesperadas = [(p, a) for p, a in sem_produto if a not in FICHAS_ESPERADAS_SEM_PRODUTO]
    return esperadas, inesperadas


# ---------------------------------------------------------------------------
# Upsert no banco (numa conexão já aberta — a transação é do chamador)
# ---------------------------------------------------------------------------


async def _upsert_conn(
    conn: Any, parsed: ParsedBase, *, remover_ausentes: bool = False
) -> tuple[dict[str, int], list[str], int]:
    """Aplica o upsert de ``parsed`` na conexão ``conn`` (o chamador controla a
    transação — a CLI abre a sua; o painel usa ``admin_connection``).

    Add/alterar = upsert por chave natural. Excluir = ``remover_ausentes``:
    produtos no banco que NÃO vieram na planilha são apagados (cascata remove
    fichas e formulações) — opt-in, nunca automático. Sem PDF, as fichas NÃO
    são tocadas (upload só de planilha não apaga fichas existentes)."""
    contagens = {"produtos": 0, "materias_primas": 0, "playbooks": 0, "formulacoes": 0, "fichas": 0}
    for p in parsed.produtos:
        await conn.execute(
            """
            INSERT INTO base_quimico_produtos
              (chave_produto, segmento, codigo_produto, nome, nome_normalizado,
               aplicacao, familia_tecnica, tipo_uso, componentes, palavras_chave,
               orientacao, atualizado_em)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9::jsonb, $10, $11, now())
            ON CONFLICT (chave_produto) DO UPDATE SET
              segmento = EXCLUDED.segmento, codigo_produto = EXCLUDED.codigo_produto,
              nome = EXCLUDED.nome, nome_normalizado = EXCLUDED.nome_normalizado,
              aplicacao = EXCLUDED.aplicacao, familia_tecnica = EXCLUDED.familia_tecnica,
              tipo_uso = EXCLUDED.tipo_uso, componentes = EXCLUDED.componentes,
              palavras_chave = EXCLUDED.palavras_chave, orientacao = EXCLUDED.orientacao,
              atualizado_em = now()
            """,
            p["chave_produto"], p["segmento"], p["codigo_produto"], p["nome"],
            p["nome_normalizado"], p["aplicacao"], p["familia_tecnica"], p["tipo_uso"],
            json.dumps(p["componentes"], ensure_ascii=False), p["palavras_chave"],
            p["orientacao"],
        )
        contagens["produtos"] += 1

    for m in parsed.mps:
        await conn.execute(
            """
            INSERT INTO base_quimico_materias_primas
              (codigo_mp, nome, formula_quimica, utilizacoes, atualizado_em)
            VALUES ($1, $2, $3, $4, now())
            ON CONFLICT (codigo_mp) DO UPDATE SET
              nome = EXCLUDED.nome, formula_quimica = EXCLUDED.formula_quimica,
              utilizacoes = EXCLUDED.utilizacoes, atualizado_em = now()
            """,
            m["codigo_mp"], m["nome"], m["formula_quimica"], m["utilizacoes"],
        )
        contagens["materias_primas"] += 1

    for pb in parsed.playbooks:
        await conn.execute(
            """
            INSERT INTO base_quimico_playbooks (tipo, sintoma, dados, atualizado_em)
            VALUES ($1, $2, $3::jsonb, now())
            ON CONFLICT (tipo, sintoma) DO UPDATE SET
              dados = EXCLUDED.dados, atualizado_em = now()
            """,
            pb["tipo"], pb["sintoma"], json.dumps(pb["dados"], ensure_ascii=False),
        )
        contagens["playbooks"] += 1

    for f in parsed.formulacoes:
        await conn.execute(
            """
            INSERT INTO base_quimico_formulacoes
              (chave_produto, ordem, codigo_mp, componente, quantidade, funcao, atualizado_em)
            VALUES ($1, $2, $3, $4, $5, $6, now())
            ON CONFLICT (chave_produto, ordem) DO UPDATE SET
              codigo_mp = EXCLUDED.codigo_mp, componente = EXCLUDED.componente,
              quantidade = EXCLUDED.quantidade, funcao = EXCLUDED.funcao,
              atualizado_em = now()
            """,
            f["chave_produto"], f["ordem"], f["codigo_mp"], f["componente"],
            f["quantidade"], f["funcao"],
        )
        contagens["formulacoes"] += 1

    # Ficha é por NOME de produto; a mesma ficha vale para toda chave (segmento)
    # que compartilha o nome.
    chaves_por_nome: dict[str, list[str]] = {}
    for p in parsed.produtos:
        chaves_por_nome.setdefault(_normalizar(p["nome"]), []).append(p["chave_produto"])
    chaves_com_ficha: list[str] = []
    for nome, conteudo in parsed.fichas_por_nome.items():
        for chave in chaves_por_nome.get(_normalizar(nome), []):
            await conn.execute(
                """
                INSERT INTO base_quimico_fichas (chave_produto, conteudo, atualizado_em)
                VALUES ($1, $2, now())
                ON CONFLICT (chave_produto) DO UPDATE SET
                  conteudo = EXCLUDED.conteudo, atualizado_em = now()
                """,
                chave, conteudo,
            )
            chaves_com_ficha.append(chave)
            contagens["fichas"] += 1
    if parsed.fichas_por_nome:
        # O PDF é a fonte da verdade das fichas: remove as órfãs de execução
        # anterior. SÓ roda quando houve PDF (upload só de planilha não mexe).
        status = await conn.execute(
            "DELETE FROM base_quimico_fichas WHERE NOT (chave_produto = ANY($1::text[]))",
            chaves_com_ficha,
        )
        contagens["fichas_removidas"] = int(status.rsplit(" ", 1)[-1])

    # Reconciliação de produtos (add/alterar já feitos; aqui é o EXCLUIR).
    chaves_planilha = [p["chave_produto"] for p in parsed.produtos]
    ausentes = [
        r["nome"]
        for r in await conn.fetch(
            "SELECT nome FROM base_quimico_produtos WHERE NOT (chave_produto = ANY($1::text[])) "
            "ORDER BY nome",
            chaves_planilha,
        )
    ]
    removidos = 0
    if remover_ausentes and ausentes:
        status = await conn.execute(
            "DELETE FROM base_quimico_produtos WHERE NOT (chave_produto = ANY($1::text[]))",
            chaves_planilha,
        )
        removidos = int(status.rsplit(" ", 1)[-1])
    return contagens, ausentes, removidos


async def ingerir_conn(
    conn: Any, planilha_bytes: bytes, pdf_bytes: bytes | None = None, *, remover_ausentes: bool = False
) -> RelatorioIngestao:
    """Parse (em thread — CPU-bound) + upsert numa conexão já transacional.

    Ponto de entrada usado pelo painel web (``admin_connection``) e reaproveitado
    pela CLI. O parse roda fora do event loop para não bloquear o servidor."""
    parsed = await asyncio.to_thread(parse_arquivos, planilha_bytes, pdf_bytes)
    contagens, ausentes, removidos = await _upsert_conn(
        conn, parsed, remover_ausentes=remover_ausentes
    )
    esperadas, inesperadas = _classificar_sem_produto(parsed.fichas_sem_produto)
    return RelatorioIngestao(contagens, esperadas, inesperadas, ausentes, removidos)
