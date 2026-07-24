"""Ingestão da base de conhecimento do Dpto Químico → tabelas ``base_quimico_*`` (F4).

plano_md_mestre_IA.md, Seção 3.3: a base (planilha de 16 abas + PDF de fichas
técnicas) é SIGILOSA e **nunca entra no repositório** — os arquivos ficam com o
gestor/dev (ex.: pasta local ou drive interno) e este script, re-executável,
lê os caminhos informados e faz **upsert** no banco a cada nova versão.

Uso (local, pelo dev/gestor — NUNCA por rota HTTP):

    python scripts/ingestao_base_quimico.py \
        --planilha "C:/caminho/Estrutura_de_Produtos_BASE_IA....xlsx" \
        --fichas-pdf "C:/caminho/Fichas Técnicas.pdf" \
        [--dsn postgresql://...]   # default: env INGESTAO_DATABASE_URL ou DATABASE_URL

Dependências extras (requirements-dev.txt): ``openpyxl`` e ``pypdf``.

O que vai para onde (Seção 3.3 do plano IA):
- ``Base_IA_Produtos``       → ``base_quimico_produtos`` (componentes SEM proporção)
- ``Base_IA_Materias_Primas``→ ``base_quimico_materias_primas``
- ``Diagnostico_Ocorrencias``→ ``base_quimico_playbooks`` (tipo DIAGNOSTICO)
- ``Perguntas_Investigacao`` → ``base_quimico_playbooks`` (tipo PERGUNTA_INVESTIGACAO)
- ``Regras_Sigilo_Resposta`` → ``base_quimico_playbooks`` (tipo REGRA_SIGILO)
- PDF de fichas (por produto)→ ``base_quimico_fichas``
- ``Base_IA_Componentes``    → ``base_quimico_formulacoes`` (quantidades — só no
  banco; sem GRANT ao role ``ia_worker``; nenhum passe de modelo lê — C7)

As abas-modelo em preenchimento (Compatibilidade_Materiais, Parametros_Controle,
RCA_6M etc.) não são ingeridas na v1 (Seção 3.3 — "não bloqueiam a v1").
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Parsing da planilha (funções puras — testáveis sem arquivo real)
# ---------------------------------------------------------------------------


def _texto(valor: Any) -> str | None:
    """Célula → str limpa (None para vazio)."""
    if valor is None:
        return None
    s = str(valor).strip()
    return s or None


def _linhas(ws) -> list[dict[str, Any]]:
    """Aba → lista de dicts {cabeçalho: valor}, ignorando linhas vazias."""
    it = ws.iter_rows(values_only=True)
    try:
        cabecalhos = [(_texto(c) or f"col{i}") for i, c in enumerate(next(it))]
    except StopIteration:
        return []
    out = []
    for row in it:
        if not any(v is not None and str(v).strip() for v in row):
            continue
        out.append({cabecalhos[i]: row[i] if i < len(row) else None for i in range(len(cabecalhos))})
    return out


def ler_produtos(ws) -> list[dict[str, Any]]:
    """Aba Base_IA_Produtos → linhas de ``base_quimico_produtos``.

    ``componentes`` vem da coluna "Componentes principais (sem proporção)" —
    NOMES separados por ';', nunca quantidades."""
    produtos = []
    for r in _linhas(ws):
        chave = _texto(r.get("Chave Produto"))
        nome = _texto(r.get("Produto"))
        if not chave or not nome:
            continue
        componentes = [
            c.strip() for c in (_texto(r.get("Componentes principais (sem proporção)")) or "").split(";") if c.strip()
        ]
        palavras = [
            p.strip() for p in (_texto(r.get("Palavras-chave para busca da IA")) or "").split() if p.strip()
        ]
        produtos.append(
            {
                "chave_produto": chave,
                "segmento": _texto(r.get("Segmento")),
                "codigo_produto": _texto(r.get("Código Produto")),
                "nome": nome,
                "nome_normalizado": _texto(r.get("Nome normalizado")),
                "aplicacao": _texto(r.get("Descrição / aplicação")),
                "familia_tecnica": _texto(r.get("Família técnica sugerida")),
                "tipo_uso": _texto(r.get("Tipo de uso sugerido")),
                "componentes": componentes,
                "palavras_chave": palavras,
                "orientacao": _texto(r.get("Orientação para resposta da IA")),
            }
        )
    return produtos


def ler_materias_primas(ws) -> list[dict[str, Any]]:
    """Aba Base_IA_Materias_Primas → linhas de ``base_quimico_materias_primas``."""
    mps = []
    for r in _linhas(ws):
        codigo = _texto(r.get("Código MP"))
        nome = _texto(r.get("Nome classificado"))
        if not codigo or not nome:
            continue
        mps.append(
            {
                "codigo_mp": codigo,
                "nome": nome,
                "formula_quimica": _texto(r.get("Fórmula química")),
                "utilizacoes": _texto(r.get("Principais utilizações")),
            }
        )
    return mps


def ler_formulacoes(ws) -> list[dict[str, Any]]:
    """Aba Base_IA_Componentes → ``base_quimico_formulacoes`` (quantidades).

    Estas linhas são "Confidencial / formulação": ficam SÓ no banco, fora do
    alcance do role ``ia_worker`` (C7) — nunca em contexto de modelo."""
    linhas = []
    for r in _linhas(ws):
        chave = _texto(r.get("Chave Produto"))
        componente = _texto(r.get("Matéria-prima / componente"))
        ordem = r.get("Ordem")
        if not chave or not componente or ordem is None:
            continue
        try:
            quantidade = float(str(r.get("Quantidade informada")).replace(",", "."))
        except (TypeError, ValueError):
            quantidade = None
        linhas.append(
            {
                "chave_produto": chave,
                "ordem": int(float(ordem)),
                "codigo_mp": _texto(r.get("Código MP/Componente")),
                "componente": componente,
                "quantidade": quantidade,
                "funcao": _texto(r.get("Função/observação extraída")),
            }
        )
    return linhas


# (tipo, coluna-chave) de cada aba de playbook — as demais colunas viram `dados`.
_PLAYBOOKS = (
    ("DIAGNOSTICO", "Diagnostico_Ocorrencias", "Sintoma / problema relatado"),
    ("PERGUNTA_INVESTIGACAO", "Perguntas_Investigacao", "Cenário / sintoma"),
    ("REGRA_SIGILO", "Regras_Sigilo_Resposta", "Tipo de informação solicitada"),
)


def ler_playbooks(wb) -> list[dict[str, Any]]:
    """Abas de diagnóstico/perguntas/regras → ``base_quimico_playbooks``."""
    out = []
    for tipo, aba, coluna_chave in _PLAYBOOKS:
        if aba not in wb.sheetnames:
            print(f"  AVISO: aba '{aba}' ausente na planilha — pulada.")
            continue
        for r in _linhas(wb[aba]):
            sintoma = _texto(r.get(coluna_chave))
            if not sintoma:
                continue
            dados = {
                k: _texto(v) for k, v in r.items() if k != coluna_chave and _texto(v) is not None
            }
            out.append({"tipo": tipo, "sintoma": sintoma, "dados": dados})
    return out


# ---------------------------------------------------------------------------
# Fatiamento do PDF de fichas por produto (função pura sobre lista de páginas)
# ---------------------------------------------------------------------------


def _normalizar(texto: str) -> str:
    """Caixa alta sem acento — comparação robusta de nomes de produto."""
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in nfkd if not unicodedata.combining(c)).upper()


def fatiar_fichas(paginas: list[str], nomes_produtos: list[str]) -> tuple[dict[str, str], list[int]]:
    """Agrupa o texto das páginas do PDF por produto citado.

    Cada página é atribuída ao produto cujo nome aparece nela (palavra inteira,
    sem acento/caixa; empate → o nome MAIS LONGO vence — evita "26" capturar
    página do "ADITIVO 1090", e o limite de palavra evita "26" casar com
    "26000 volts"). Página sem produto identificado é tratada como CONTINUAÇÃO
    da ficha anterior; se não houver anterior, fica sem dona (relatada para
    revisão manual).

    Retorna ``({nome_produto: texto_concatenado}, [páginas_sem_dona_1based])``.
    """
    padroes = sorted(
        ((nome, re.compile(rf"(?<![A-Z0-9]){re.escape(_normalizar(nome))}(?![A-Z0-9])")) for nome in nomes_produtos),
        key=lambda par: len(par[0]),
        reverse=True,
    )
    fichas: dict[str, list[str]] = {}
    sem_dona: list[int] = []
    atual: str | None = None
    for i, texto in enumerate(paginas):
        norm = _normalizar(texto or "")
        dono = next((nome for nome, padrao in padroes if padrao.search(norm)), None)
        if dono is None:
            if atual is None:
                sem_dona.append(i + 1)
                continue
            dono = atual  # continuação da ficha anterior
        atual = dono
        fichas.setdefault(dono, []).append((texto or "").strip())
    return {nome: "\n\n".join(partes) for nome, partes in fichas.items()}, sem_dona


# ---------------------------------------------------------------------------
# Upsert no banco (asyncpg; transação única — tudo ou nada)
# ---------------------------------------------------------------------------


async def _upsert_tudo(
    dsn: str,
    produtos: list[dict[str, Any]],
    mps: list[dict[str, Any]],
    playbooks: list[dict[str, Any]],
    formulacoes: list[dict[str, Any]],
    fichas_por_nome: dict[str, str],
) -> dict[str, int]:
    import asyncpg

    conn = await asyncpg.connect(dsn, statement_cache_size=0)
    contagens = {"produtos": 0, "materias_primas": 0, "playbooks": 0, "formulacoes": 0, "fichas": 0}
    try:
        async with conn.transaction():
            for p in produtos:
                await conn.execute(
                    """
                    INSERT INTO base_quimico_produtos
                      (chave_produto, segmento, codigo_produto, nome, nome_normalizado,
                       aplicacao, familia_tecnica, tipo_uso, componentes, palavras_chave,
                       orientacao, atualizado_em)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9::jsonb, $10, $11, now())
                    ON CONFLICT (chave_produto) DO UPDATE SET
                      segmento = EXCLUDED.segmento, codigo_produto = EXCLUDED.codigo_produto,
                      nome = EXCLUDED.nome, nome_normalizado = EXCLUDED.nome_normalizado,
                      aplicacao = EXCLUDED.aplicacao, familia_tecnica = EXCLUDED.familia_tecnica,
                      tipo_uso = EXCLUDED.tipo_uso, componentes = EXCLUDED.componentes,
                      palavras_chave = EXCLUDED.palavras_chave, orientacao = EXCLUDED.orientacao,
                      atualizado_em = now()
                    """,
                    p["chave_produto"], p["segmento"], p["codigo_produto"], p["nome"],
                    p["nome_normalizado"], p["aplicacao"], p["familia_tecnica"], p["tipo_uso"],
                    json.dumps(p["componentes"], ensure_ascii=False), p["palavras_chave"],
                    p["orientacao"],
                )
                contagens["produtos"] += 1

            for m in mps:
                await conn.execute(
                    """
                    INSERT INTO base_quimico_materias_primas
                      (codigo_mp, nome, formula_quimica, utilizacoes, atualizado_em)
                    VALUES ($1, $2, $3, $4, now())
                    ON CONFLICT (codigo_mp) DO UPDATE SET
                      nome = EXCLUDED.nome, formula_quimica = EXCLUDED.formula_quimica,
                      utilizacoes = EXCLUDED.utilizacoes, atualizado_em = now()
                    """,
                    m["codigo_mp"], m["nome"], m["formula_quimica"], m["utilizacoes"],
                )
                contagens["materias_primas"] += 1

            for pb in playbooks:
                await conn.execute(
                    """
                    INSERT INTO base_quimico_playbooks (tipo, sintoma, dados, atualizado_em)
                    VALUES ($1, $2, $3::jsonb, now())
                    ON CONFLICT (tipo, sintoma) DO UPDATE SET
                      dados = EXCLUDED.dados, atualizado_em = now()
                    """,
                    pb["tipo"], pb["sintoma"], json.dumps(pb["dados"], ensure_ascii=False),
                )
                contagens["playbooks"] += 1

            for f in formulacoes:
                await conn.execute(
                    """
                    INSERT INTO base_quimico_formulacoes
                      (chave_produto, ordem, codigo_mp, componente, quantidade, funcao, atualizado_em)
                    VALUES ($1, $2, $3, $4, $5, $6, now())
                    ON CONFLICT (chave_produto, ordem) DO UPDATE SET
                      codigo_mp = EXCLUDED.codigo_mp, componente = EXCLUDED.componente,
                      quantidade = EXCLUDED.quantidade, funcao = EXCLUDED.funcao,
                      atualizado_em = now()
                    """,
                    f["chave_produto"], f["ordem"], f["codigo_mp"], f["componente"],
                    f["quantidade"], f["funcao"],
                )
                contagens["formulacoes"] += 1

            # Ficha é por NOME de produto; a mesma ficha vale para toda chave
            # (segmento) que compartilha o nome.
            chaves_por_nome: dict[str, list[str]] = {}
            for p in produtos:
                chaves_por_nome.setdefault(_normalizar(p["nome"]), []).append(p["chave_produto"])
            for nome, conteudo in fichas_por_nome.items():
                for chave in chaves_por_nome.get(_normalizar(nome), []):
                    await conn.execute(
                        """
                        INSERT INTO base_quimico_fichas (chave_produto, conteudo, atualizado_em)
                        VALUES ($1, $2, now())
                        ON CONFLICT (chave_produto) DO UPDATE SET
                          conteudo = EXCLUDED.conteudo, atualizado_em = now()
                        """,
                        chave, conteudo,
                    )
                    contagens["fichas"] += 1
    finally:
        await conn.close()
    return contagens


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.split("\n", 1)[0])
    parser.add_argument("--planilha", required=True, help="Caminho do .xlsx (fora do repo)")
    parser.add_argument("--fichas-pdf", help="Caminho do PDF de fichas técnicas (fora do repo)")
    parser.add_argument(
        "--dsn",
        default=os.environ.get("INGESTAO_DATABASE_URL") or os.environ.get("DATABASE_URL"),
        help="DSN administrativa do Postgres (default: INGESTAO_DATABASE_URL ou DATABASE_URL)",
    )
    args = parser.parse_args(argv)
    if not args.dsn:
        parser.error("--dsn ausente e nem INGESTAO_DATABASE_URL/DATABASE_URL definidas.")

    planilha = Path(args.planilha)
    if not planilha.is_file():
        parser.error(f"Planilha não encontrada: {planilha}")

    import openpyxl

    print(f"Lendo planilha: {planilha.name}")
    wb = openpyxl.load_workbook(planilha, read_only=True, data_only=True)
    produtos = ler_produtos(wb["Base_IA_Produtos"])
    mps = ler_materias_primas(wb["Base_IA_Materias_Primas"])
    formulacoes = ler_formulacoes(wb["Base_IA_Componentes"])
    playbooks = ler_playbooks(wb)

    fichas_por_nome: dict[str, str] = {}
    sem_dona: list[int] = []
    if args.fichas_pdf:
        fichas_pdf = Path(args.fichas_pdf)
        if not fichas_pdf.is_file():
            parser.error(f"PDF não encontrado: {fichas_pdf}")
        from pypdf import PdfReader

        print(f"Lendo fichas: {fichas_pdf.name}")
        reader = PdfReader(str(fichas_pdf))
        paginas = [(page.extract_text() or "") for page in reader.pages]
        fichas_por_nome, sem_dona = fatiar_fichas(paginas, [p["nome"] for p in produtos])

    print(
        f"Planilha: {len(produtos)} produtos, {len(mps)} matérias-primas, "
        f"{len(playbooks)} playbooks, {len(formulacoes)} linhas de formulação (confidenciais)."
    )
    if args.fichas_pdf:
        print(f"PDF: fichas identificadas para {len(fichas_por_nome)} produtos.")
        if sem_dona:
            print(f"  AVISO: páginas sem produto identificado (revisar manualmente): {sem_dona}")

    contagens = asyncio.run(
        _upsert_tudo(args.dsn, produtos, mps, playbooks, formulacoes, fichas_por_nome)
    )
    print("Upsert concluído:", ", ".join(f"{k}={v}" for k, v in contagens.items()))
    print("Re-execução é segura: mesma chave natural ⇒ UPDATE, nunca duplica.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
