"""Leitura da planilha "Investimento por Região" (`app/services/ingestao_marketing_midia.py`).

Layout real validado à mão contra a planilha de referência do usuário
(2026-07-30): cada aba é um mês ("Março 2026"), coluna A marca representante
ativo com preenchimento verde sólido, a coluna "5.Valor mensal do
investimento (vendedor + Bondmann):" é o investimento real e a coluna
"6.Observações" registra entrada/saída com o template "Entrou em .../Saiu
em ...". Usa `openpyxl.Workbook()` de verdade (não um fake) porque a lógica
depende de estilo de célula (`cell.fill`), que um fake baseado em
`iter_rows(values_only=True)` não consegue simular.
"""

from __future__ import annotations

import openpyxl
from openpyxl.styles import PatternFill

from app.services.ingestao_marketing_midia import ler_aba, mes_da_aba, parse_planilha

_VERDE = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")

_CABECALHO_NOVO = (
    None, "Representante", "Valor mensal do investimento disponível (vendedor + Bondmann)",
    "Produtos Divulgados", "Número de Usuários Alcançados", "Número de Cliques no Anúncio",
    "Custo por Clique nos Anúncios", "Nº de Leads Identificados", "Custo por Leads Identificados",
    "5.Valor mensal do investimento (vendedor + Bondmann):", "6.Observações", "Valor mensal Neat",
)

_CABECALHO_ANTIGO = (
    None, "Representante", "Valor mensal do investimento disponível (vendedor + Bondmann)",
    "Número de Usuários Alcançados", "Número de Cliques no Anúncio", "Custo por Clique nos Anúncios",
    "5.Valor mensal do investimento (vendedor + Bondmann):",
    "6.Dados time comercial: Nº mensagens recebidas", "7.Comentários time comercial",
)


def _aba_mes(wb, nome: str, cabecalho: tuple, linhas: list[tuple], *, verdes: set[int]):
    """Cria uma aba no layout real: cabeçalho na linha 1, dados a partir da
    linha 2, linha "TOTAL" ao final. ``verdes`` = números da coluna A (1-based
    dentro das linhas) que recebem o preenchimento verde de "ativo"."""
    ws = wb.create_sheet(nome)
    for c, texto in enumerate(cabecalho, start=1):
        ws.cell(row=1, column=c, value=texto)
    for i, linha in enumerate(linhas, start=1):
        r = i + 1
        for c, valor in enumerate(linha, start=1):
            ws.cell(row=r, column=c, value=valor)
        if i in verdes:
            ws.cell(row=r, column=1).fill = _VERDE
    ws.cell(row=len(linhas) + 2, column=1, value="TOTAL")
    return ws


def test_mes_da_aba_reconhece_mes_em_portugues_e_ignora_o_resto():
    assert mes_da_aba("Março 2026").isoformat() == "2026-03-01"
    assert mes_da_aba("Junho 2025").isoformat() == "2025-06-01"
    assert mes_da_aba("Copy of Fevereiro 2025") is None
    assert mes_da_aba("Indicadores") is None
    assert mes_da_aba("Gráficos") is None


def test_ler_aba_layout_novo_investimento_regioes_e_observacoes():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    linhas = [
        (1, "Rep 1", 200.0, "X", 100, "10", 1.0, 5, 2.0, 200.02, None, 100.0),
        (2, "Rep 2", 100.0, "X", 100, "10", 1.0, 5, 2.0, 99.93, "Entrou em 9 de Junho.", 100.0),
        (3, "Rep 3", 100.0, "X", 100, "10", 1.0, 5, 2.0, 99.99, None, 100.0),
        # Rep 4 não é verde (não conta pra "regiões ativas") mas ainda soma no investimento.
        (4, "Rep 4 (inativo)", 100.0, "X", 100, "10", 1.0, 5, 2.0, 50.0, "Saiu em 25 de Junho", 100.0),
    ]
    ws = _aba_mes(wb, "Junho 2026", _CABECALHO_NOVO, linhas, verdes={1, 2, 3})

    resultado = ler_aba(ws)
    assert resultado is not None
    investimento, regioes, descontinuidades, aderencias = resultado
    # (200.02 + 99.93 + 99.99 + 50.0) / 2 = 224.97
    assert investimento == 224.97
    assert regioes == 3  # maior número da coluna A com preenchimento verde
    assert descontinuidades == 1
    assert aderencias == 1


def test_ler_aba_layout_antigo_sem_coluna_de_observacoes_zera_entradas_saidas():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    linhas = [
        (1, "Rep 1", 400.0, 14135, "513", 0.66, 337.6, None, None),
        (2, "Rep 2", 200.0, 15896, "471", 0.42, 200.01, None, None),
    ]
    ws = _aba_mes(wb, "Novembro 2024", _CABECALHO_ANTIGO, linhas, verdes={1, 2})

    investimento, regioes, descontinuidades, aderencias = ler_aba(ws)
    assert investimento == round((337.6 + 200.01) / 2, 2)
    assert regioes == 2
    # Layout antigo não tem coluna "6.Observações" — nunca inventa dado.
    assert descontinuidades == 0
    assert aderencias == 0


def test_ler_aba_ignora_observacoes_fora_do_template_combinado():
    """Só "Saiu"/"Entrou" no início do texto contam — texto livre antigo tipo
    "Retornou em Janeiro"/"Promovido a supervisor" fica de fora de propósito
    (nunca adivinha sinônimo do template combinado com o usuário)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    linhas = [
        (1, "Rep 1", 100.0, "X", 1, "1", 1.0, 1, 1.0, 100.0, "Início em 12 de Janeiro", 100.0),
        (2, "Rep 2", 100.0, "X", 1, "1", 1.0, 1, 1.0, 100.0, "Retornou em Janeiro", 100.0),
        (3, "Rep 3", 100.0, "X", 1, "1", 1.0, 1, 1.0, 100.0, "Saiu em 22 de Janeiro", 100.0),
    ]
    ws = _aba_mes(wb, "Janeiro 2026", _CABECALHO_NOVO, linhas, verdes={1, 2, 3})

    _investimento, _regioes, descontinuidades, aderencias = ler_aba(ws)
    assert descontinuidades == 1
    assert aderencias == 0


def test_parse_planilha_pula_abas_copy_of_e_abas_sem_cabecalho_esperado():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    linhas = [(1, "Rep 1", 200.0, "X", 1, "1", 1.0, 1, 1.0, 200.0, None, 100.0)]
    _aba_mes(wb, "Março 2026", _CABECALHO_NOVO, linhas, verdes={1})
    _aba_mes(wb, "Copy of Março 2026", _CABECALHO_NOVO, linhas, verdes={1})
    resumo = wb.create_sheet("Indicadores")
    resumo.cell(row=1, column=1, value="Resumo — não é aba de mês")

    registros = parse_planilha(wb)
    assert [r.mes.isoformat() for r in registros] == ["2026-03-01"]


def test_parse_planilha_ordena_por_mes():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    linhas = [(1, "Rep 1", 100.0, "X", 1, "1", 1.0, 1, 1.0, 100.0, None, 100.0)]
    _aba_mes(wb, "Junho 2026", _CABECALHO_NOVO, linhas, verdes={1})
    _aba_mes(wb, "Janeiro 2026", _CABECALHO_NOVO, linhas, verdes={1})
    _aba_mes(wb, "Março 2026", _CABECALHO_NOVO, linhas, verdes={1})

    registros = parse_planilha(wb)
    assert [r.mes.isoformat() for r in registros] == ["2026-01-01", "2026-03-01", "2026-06-01"]
