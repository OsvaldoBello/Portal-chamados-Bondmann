"""Leitura da planilha da base do Químico (`app/services/ingestao_quimico.py`).

Cobre a parte determinística e sem I/O da ingestão — as abas viram linhas de
tabela — com um workbook falso (uma aba é só uma lista de tuplas, que é tudo
que `openpyxl` entrega com `values_only=True`). O fatiamento do PDF de fichas
já tem cobertura própria em `tests/test_ia_quimico.py`; aqui o foco é o parse
da planilha e as invariantes que protegem o sigilo:

- **quantidades** só existem em `ler_formulacoes` (tabela que o role `ia_worker`
  não alcança — C7); `ler_produtos` traz NOMES de componentes, nunca proporção;
- linha sem chave/nome é descartada em silêncio, e não vira registro capenga.
"""

from __future__ import annotations

from app.services.ingestao_quimico import (
    _classificar_sem_produto,
    ler_formulacoes,
    ler_materias_primas,
    ler_playbooks,
    ler_produtos,
    parse_planilha,
)


class _FakeWs:
    """Aba de planilha: cabeçalho + linhas, como `openpyxl` com values_only."""

    def __init__(self, linhas: list[tuple]):
        self._linhas = linhas

    def iter_rows(self, values_only: bool = True):
        yield from self._linhas


class _FakeWb:
    def __init__(self, abas: dict[str, _FakeWs]):
        self._abas = abas
        self.sheetnames = list(abas)

    def __getitem__(self, nome: str) -> _FakeWs:
        return self._abas[nome]


def _ws_produtos(*linhas: tuple) -> _FakeWs:
    cabecalho = (
        "Chave Produto", "Segmento", "Código Produto", "Produto", "Nome normalizado",
        "Descrição / aplicação", "Família técnica sugerida", "Tipo de uso sugerido",
        "Componentes principais (sem proporção)", "Palavras-chave para busca da IA",
        "Orientação para resposta da IA",
    )
    return _FakeWs([cabecalho, *linhas])


def test_aba_vazia_nao_quebra():
    assert ler_produtos(_FakeWs([])) == []


def test_ler_produtos_quebra_componentes_e_palavras_chave():
    ws = _ws_produtos(
        ("P1", "Automotivo", "COD1", "BRIL", "bril", "Limpeza", "Detergente", "Diluído",
         "Tensoativo aniônico; Sequestrante ; ", "bril brilho lavagem", "Citar ficha"),
    )
    (produto,) = ler_produtos(ws)
    assert produto["chave_produto"] == "P1"
    assert produto["nome"] == "BRIL"
    # ';' separa componentes (nomes, sem proporção); espaço separa palavras-chave.
    assert produto["componentes"] == ["Tensoativo aniônico", "Sequestrante"]
    assert produto["palavras_chave"] == ["bril", "brilho", "lavagem"]
    assert produto["segmento"] == "Automotivo"


def test_ler_produtos_descarta_linha_sem_chave_ou_sem_nome():
    ws = _ws_produtos(
        (None, "Automotivo", "COD1", "SEM CHAVE", None, None, None, None, None, None, None),
        ("P2", "Automotivo", "COD2", "   ", None, None, None, None, None, None, None),
        (None, None, None, None, None, None, None, None, None, None, None),  # linha vazia
    )
    assert ler_produtos(ws) == []


def test_ler_materias_primas_exige_codigo_e_nome():
    ws = _FakeWs([
        ("Código MP", "Nome classificado", "Fórmula química", "Principais utilizações"),
        ("MP1", "Soda cáustica", "NaOH", "Alcalinizante"),
        ("MP2", None, "H2O", "Diluente"),  # sem nome → fora
    ])
    (mp,) = ler_materias_primas(ws)
    assert mp == {
        "codigo_mp": "MP1", "nome": "Soda cáustica",
        "formula_quimica": "NaOH", "utilizacoes": "Alcalinizante",
    }


def test_ler_formulacoes_aceita_virgula_decimal_e_ignora_quantidade_invalida():
    ws = _FakeWs([
        ("Chave Produto", "Ordem", "Código MP/Componente", "Matéria-prima / componente",
         "Quantidade informada", "Função/observação extraída"),
        ("P1", 1, "MP1", "Soda cáustica", "12,5", "Alcalinizante"),
        ("P1", 2.0, "MP2", "Água", "qs", None),          # quantidade não numérica
        ("P1", None, "MP3", "Corante", "1", None),        # sem ordem → fora
        (None, 4, "MP4", "Perfume", "1", None),           # sem chave → fora
    ])
    linhas = ler_formulacoes(ws)
    assert [linha["ordem"] for linha in linhas] == [1, 2]
    assert linhas[0]["quantidade"] == 12.5
    assert linhas[1]["quantidade"] is None  # "qs" não vira 0 nem quebra a ingestão


def test_ler_playbooks_pula_aba_ausente_e_move_o_resto_para_dados():
    wb = _FakeWb({
        "Perguntas_Investigacao": _FakeWs([
            ("Cenário / sintoma", "Pergunta", "Observação"),
            ("Excesso de espuma", "Qual a dosagem usada?", None),
            (None, "órfã sem cenário", "x"),
        ]),
        # Diagnostico_Ocorrencias e Regras_Sigilo_Resposta ausentes de propósito.
    })
    (item,) = ler_playbooks(wb)
    assert item["tipo"] == "PERGUNTA_INVESTIGACAO"
    assert item["sintoma"] == "Excesso de espuma"
    # A coluna-chave não se repete dentro de `dados`; colunas vazias somem.
    assert item["dados"] == {"Pergunta": "Qual a dosagem usada?"}


def test_parse_planilha_junta_produtos_so_ficha():
    """Os produtos comprados prontos (AW-B/LW-B) não estão na planilha — entram
    pelo catálogo em código para que a ficha do PDF tenha onde ancorar."""
    wb = _FakeWb({
        "Base_IA_Produtos": _ws_produtos(
            ("P1", "Automotivo", "COD1", "BRIL", None, None, None, None, None, None, None),
        ),
        "Base_IA_Materias_Primas": _FakeWs([("Código MP", "Nome classificado")]),
        "Base_IA_Componentes": _FakeWs([("Chave Produto", "Ordem")]),
    })
    produtos, mps, formulacoes, playbooks = parse_planilha(wb)
    nomes = [p["nome"] for p in produtos]
    assert "BRIL" in nomes
    assert len(nomes) > 1  # produtos-só-ficha entraram junto
    assert (mps, formulacoes, playbooks) == ([], [], [])


class _FakeConn:
    """Conexão asyncpg mínima: grava o SQL e devolve o combinado."""

    def __init__(self, ausentes: list[str] | None = None, apagados: int = 0):
        self._ausentes = ausentes or []
        self._apagados = apagados
        self.sqls: list[str] = []

    async def execute(self, sql: str, *args) -> str:
        self.sqls.append(" ".join(sql.split()))
        return f"DELETE {self._apagados}"

    async def fetch(self, sql: str, *args) -> list[dict]:
        self.sqls.append(" ".join(sql.split()))
        return [{"nome": nome} for nome in self._ausentes]

    def _tabelas_escritas(self) -> set[str]:
        import re as _re

        return {m.group(1) for s in self.sqls for m in [_re.search(r"INTO (\w+)", s)] if m}


def _parsed(**over):
    from app.services.ingestao_quimico import ParsedBase

    base = {
        "produtos": [{
            "chave_produto": "P1", "segmento": "Automotivo", "codigo_produto": "C1",
            "nome": "BRIL", "nome_normalizado": "bril", "aplicacao": None,
            "familia_tecnica": None, "tipo_uso": None, "componentes": ["Tensoativo"],
            "palavras_chave": ["bril"], "orientacao": None,
        }],
        "mps": [{"codigo_mp": "MP1", "nome": "Soda", "formula_quimica": None,
                 "utilizacoes": None}],
        "playbooks": [{"tipo": "DIAGNOSTICO", "sintoma": "Espuma", "dados": {"a": "b"}}],
        "formulacoes": [{"chave_produto": "P1", "ordem": 1, "codigo_mp": "MP1",
                         "componente": "Soda", "quantidade": 12.5, "funcao": None}],
        "fichas_por_nome": {},
        "fichas_sem_produto": [],
    }
    base.update(over)
    return ParsedBase(**base)


async def test_upsert_conta_cada_tabela_e_nao_toca_fichas_sem_pdf():
    """Upload só de planilha NÃO pode apagar as fichas já ingeridas do PDF."""
    from app.services.ingestao_quimico import _upsert_conn

    conn = _FakeConn()
    contagens, ausentes, removidos = await _upsert_conn(conn, _parsed())
    assert contagens == {"produtos": 1, "materias_primas": 1, "playbooks": 1,
                         "formulacoes": 1, "fichas": 0}
    assert (ausentes, removidos) == ([], 0)
    assert "base_quimico_fichas" not in conn._tabelas_escritas()
    assert not any("DELETE FROM base_quimico_fichas" in s for s in conn.sqls)


async def test_upsert_com_pdf_grava_ficha_por_nome_e_limpa_orfas():
    from app.services.ingestao_quimico import _upsert_conn

    conn = _FakeConn(apagados=2)
    # O nome da ficha casa com o do produto por forma normalizada (acento/caixa).
    contagens, _, _ = await _upsert_conn(
        conn, _parsed(fichas_por_nome={"bril": "ficha do BRIL", "SUMIU": "órfã"})
    )
    assert contagens["fichas"] == 1  # só a que tem produto correspondente
    assert contagens["fichas_removidas"] == 2
    assert any("DELETE FROM base_quimico_fichas" in s for s in conn.sqls)


async def test_remover_ausentes_e_opt_in():
    from app.services.ingestao_quimico import _upsert_conn

    conn = _FakeConn(ausentes=["PRODUTO ANTIGO"], apagados=1)
    _, ausentes, removidos = await _upsert_conn(conn, _parsed())
    assert ausentes == ["PRODUTO ANTIGO"]
    assert removidos == 0  # sem a flag, só REPORTA quem sumiu da planilha
    assert not any("DELETE FROM base_quimico_produtos" in s for s in conn.sqls)

    conn = _FakeConn(ausentes=["PRODUTO ANTIGO"], apagados=1)
    _, _, removidos = await _upsert_conn(conn, _parsed(), remover_ausentes=True)
    assert removidos == 1
    assert any("DELETE FROM base_quimico_produtos" in s for s in conn.sqls)


def _planilha_bytes() -> bytes:
    """Planilha .xlsx de verdade (openpyxl é dependência de produção), com as 3
    abas obrigatórias — é o formato que o painel `/admin/base-quimico` recebe."""
    from io import BytesIO

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Base_IA_Produtos"
    ws.append([
        "Chave Produto", "Segmento", "Código Produto", "Produto", "Nome normalizado",
        "Descrição / aplicação", "Família técnica sugerida", "Tipo de uso sugerido",
        "Componentes principais (sem proporção)", "Palavras-chave para busca da IA",
        "Orientação para resposta da IA",
    ])
    ws.append(["P1", "Automotivo", "C1", "BRIL", "bril", None, None, None,
               "Tensoativo", "bril", None])
    wb.create_sheet("Base_IA_Materias_Primas").append(["Código MP", "Nome classificado"])
    wb.create_sheet("Base_IA_Componentes").append(["Chave Produto", "Ordem"])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_arquivos_le_xlsx_real_sem_pdf():
    from app.services.ingestao_quimico import parse_arquivos

    parsed = parse_arquivos(_planilha_bytes())
    assert "BRIL" in [p["nome"] for p in parsed.produtos]
    # Sem PDF não há ficha nenhuma — nem vazia, nem "sem produto".
    assert parsed.fichas_por_nome == {}
    assert parsed.fichas_sem_produto == []


def test_parse_arquivos_com_pdf_fatia_as_fichas(monkeypatch):
    """O PDF entra como lista de páginas de texto; o fatiamento em si tem
    cobertura dedicada em `test_ia_quimico.py` — aqui só se prova a costura."""

    class _FakePage:
        def __init__(self, texto):
            self._texto = texto

        def extract_text(self):
            return self._texto

    class _FakeReader:
        def __init__(self, _buf):
            self.pages = [_FakePage(">> BRIL\nFicha técnica do produto.")]

    monkeypatch.setattr("pypdf.PdfReader", _FakeReader)

    from app.services.ingestao_quimico import parse_arquivos

    parsed = parse_arquivos(_planilha_bytes(), b"%PDF-1.4 nao importa")
    assert "BRIL" in parsed.fichas_por_nome


async def test_ingerir_conn_devolve_relatorio_completo(monkeypatch):
    """Ponto de entrada do painel: parse (fora do event loop) + upsert + a
    classificação das fichas sem produto viram um `RelatorioIngestao`."""
    from app.services import ingestao_quimico as mod

    conhecida = next(iter(mod.FICHAS_ESPERADAS_SEM_PRODUTO))
    monkeypatch.setattr(
        mod, "parse_arquivos",
        lambda *a, **k: _parsed(fichas_sem_produto=[(3, conhecida), (9, "FICHA NOVA")]),
    )
    conn = _FakeConn(ausentes=["SUMIU"])
    relatorio = await mod.ingerir_conn(conn, b"planilha", None)
    assert relatorio.contagens["produtos"] == 1
    assert relatorio.fichas_esperadas == [(3, conhecida, mod.FICHAS_ESPERADAS_SEM_PRODUTO[conhecida])]
    assert relatorio.fichas_inesperadas == [(9, "FICHA NOVA")]
    assert relatorio.produtos_ausentes == ["SUMIU"]
    assert relatorio.produtos_removidos == 0


def test_classificar_sem_produto_separa_conhecidas_de_inesperadas():
    from app.services.ingestao_quimico import FICHAS_ESPERADAS_SEM_PRODUTO

    conhecida = next(iter(FICHAS_ESPERADAS_SEM_PRODUTO))
    esperadas, inesperadas = _classificar_sem_produto([(3, conhecida), (7, "FICHA NOVA")])
    assert esperadas == [(3, conhecida, FICHAS_ESPERADAS_SEM_PRODUTO[conhecida])]
    assert inesperadas == [(7, "FICHA NOVA")]
