"""Exportação organizada/dinâmica do Dashboard de Marketing
(`app/services/export_marketing.py`).

Cobre o filtro por período (mesmo contrato do `data-filter` dos "pills" da
tela: "all" ou o rótulo de um mês) e a montagem das 5 abas — os valores
precisam bater com a mesma agregação que `admin_marketing.js` faz no
navegador (soma/pct/ranking), já que a planilha é gerada a partir do MESMO
dict (`AdminRepo.mkt_dashboard_data`), sem segunda fonte de verdade.
"""

from __future__ import annotations

import openpyxl

from app.services.export_marketing import gerar_workbook

_MKT_DATA = {
    "monthly": [
        {
            "label": "JAN/26", "total": 2, "concluidas": 0, "em_andamento": 0, "abertas": 0,
            "volume": 6, "mkt_orig": 0, "sol_orig": 2, "tempo_medio": 184.2, "atrasos": 2,
            "pct_conc": 0.0, "pct_mkt": 0.0,
        },
        {
            "label": "JUL/26", "total": 76, "concluidas": 49, "em_andamento": 18, "abertas": 16,
            "volume": 314, "mkt_orig": 1, "sol_orig": 75, "tempo_medio": 4.9, "atrasos": 32,
            "pct_conc": 64.5, "pct_mkt": 1.3,
        },
    ],
    "deptByMonth": {
        "JAN/26": {"Marketing": 2},
        "JUL/26": {"Marketing": 51, "TI": 10, "RH": 15},
    },
    "atrasosData": [
        {"nome": "Chamado A", "mes": "JAN/26", "dias": 184, "causa": "Sem causa registrada"},
        {"nome": "Chamado B", "mes": "JUL/26", "dias": 7, "causa": "Aguardando definição interna"},
    ],
    "midia": {
        "meses": ["Jan", "Jul"],
        "investimento": [4605.02, 5000.0],
        "regioes": [51, 60],
        "descontinuidades": [7, 2],
        "aderencias": [2, 4],
    },
}


def _sheet_rows(wb, nome: str) -> list[tuple]:
    ws = wb[nome]
    return [row for row in ws.iter_rows(values_only=True) if any(v is not None for v in row)]


def test_abas_esperadas_na_ordem():
    wb = openpyxl.load_workbook(gerar_workbook(_MKT_DATA, "all"))
    assert wb.sheetnames == [
        "Resumo", "Volume Mensal", "Solicitantes", "Tempo e Atrasos", "Mídia Regional",
    ]


def test_acumulado_agrega_todos_os_meses():
    wb = openpyxl.load_workbook(gerar_workbook(_MKT_DATA, "all"))
    resumo = dict(_sheet_rows(wb, "Resumo")[2:])  # pula título + cabeçalho
    assert resumo["Total de demandas"] == 78  # 2 + 76
    assert resumo["Concluídas"] == 49
    assert resumo["Atrasos > 5 dias"] == 2  # len(atrasosData) sem filtro
    assert resumo["Maior setor solicitante"] == "Marketing"
    assert resumo["Demandas do maior solicitante"] == 53  # 2 + 51

    volume_rows = _sheet_rows(wb, "Volume Mensal")[2:]
    assert [r[0] for r in volume_rows] == ["JAN/26", "JUL/26"]

    solicitantes = _sheet_rows(wb, "Solicitantes")[2:]
    assert solicitantes == [("Marketing", 53), ("RH", 15), ("TI", 10)]  # ordenado desc

    midia_rows = _sheet_rows(wb, "Mídia Regional")[2:]
    assert [r[0] for r in midia_rows] == ["Jan", "Jul"]


def test_filtro_por_mes_restringe_todas_as_abas():
    wb = openpyxl.load_workbook(gerar_workbook(_MKT_DATA, "JUL/26"))
    resumo = dict(_sheet_rows(wb, "Resumo")[2:])
    assert resumo["Total de demandas"] == 76
    assert resumo["Atrasos > 5 dias"] == 1  # só o atraso de JUL/26
    assert resumo["Demandas do maior solicitante"] == 51  # não soma JAN/26

    volume_rows = _sheet_rows(wb, "Volume Mensal")[2:]
    assert [r[0] for r in volume_rows] == ["JUL/26"]

    atrasos_rows = _sheet_rows(wb, "Tempo e Atrasos")[2:]
    assert [r[0] for r in atrasos_rows] == ["Chamado B"]  # Chamado A é de JAN/26

    midia_rows = _sheet_rows(wb, "Mídia Regional")[2:]
    assert midia_rows == [("Jul", 5000.0, 60, 2, 4)]


def test_titulo_da_aba_resumo_reflete_o_periodo():
    wb_all = openpyxl.load_workbook(gerar_workbook(_MKT_DATA, "all"))
    assert wb_all["Resumo"]["A1"].value == "Resumo — Acumulado"

    wb_mes = openpyxl.load_workbook(gerar_workbook(_MKT_DATA, "JAN/26"))
    assert wb_mes["Resumo"]["A1"].value == "Resumo — JAN/26"


def test_periodo_sem_dados_gera_planilha_vazia_sem_lancar():
    wb = openpyxl.load_workbook(gerar_workbook(_MKT_DATA, "DEZ/99"))
    resumo = dict(_sheet_rows(wb, "Resumo")[2:])
    assert resumo["Total de demandas"] == 0
    assert resumo["Maior setor solicitante"] == "—"
